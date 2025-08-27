from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import MergedCell

from .File import File
from pathlib import Path

from ..Models.Fileschame import *

from .JsonFile import JsonFile

from re import findall
from io import BytesIO
from copy import deepcopy, copy


class XlsxFile(File):
    def __init__(self, map_data: FileSchemas, content: bytes | None = None):
        self.__buffer: BytesIO = BytesIO(content)
        self._file: Workbook = Workbook()
        self.__map_data: FileSchemas = map_data
        self.__active_sheet: Worksheet = None

        self.__border_style: Border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
        self.__fill_yellow: PatternFill = PatternFill(start_color="5cb800",
                                                      fill_type="solid")

    @property
    def file(self):
        return self._file

    @property
    def map_data(self):
        return self.__map_data

    @property
    def active_sheet(self):
        return self.__active_sheet

    def read_file(self, data_only: bool = True):
        self._file = load_workbook(self.__buffer, data_only=data_only)

    def render(self, schemas: dict):
        pass

    def is_table_schema(self, table, re: str):
        pass

    def get_all_parser_table_in_file(self) -> list[TableSchemas]:
        list_tables = []
        for protocols in self.__map_data.protocols:
            list_tables += protocols.tables
        return list_tables

    def get_list_cells(self, table):
        pass

    def get_sheet_by_name(self, name: str):
        return self._file.get_sheet_by_name(name)

    def target_sheet_by_name(self, name: str):
        self.__active_sheet = self._file.get_sheet_by_name(name)

    def create_sheet(self, protocol: Protocol):
        self.__active_sheet = self._file.create_sheet(protocol.name)

    def get_cell(self, x: int, y: int):
        return self.__active_sheet.cell(row=y, column=x)

    def create_cell(self, x: int, y: int, text: str):
        try:
            object_cell = self.get_cell(x, y)
            object_cell.value = text
            object_cell.border = self.__border_style
        except:
            pass

    def get_coord_cell(self, x: int, y: int) -> str:
        cell = self.get_cell(x, y)
        column_letter = cell.column_letter
        row = cell.row
        return f"{column_letter}{row}"

    def add_fill_cell(self, x: int, y: int):
        cell = self.get_cell(x, y)
        cell.fill = self.__fill_yellow

    def merge_cells(self, start_x: int, start_y: int, end_x: int, end_y: int):
        self.__active_sheet.merge_cells(start_row=start_y,
                                        start_column=start_x,
                                        end_row=end_y,
                                        end_column=end_x)

    def get_title_sheet(self) -> str:
        return self.__active_sheet.title

    def get_text_in_cell(self, text) -> str:
        text_list = findall(r"\w+", text)
        return text_list[0]

    def copy_sheet(self, name_sheet: str, target_file):

        self.target_sheet_by_name(name_sheet)
        target_file.target_sheet_by_name(name_sheet)
        self.__copy_cells(self.__active_sheet, target_file.active_sheet)  # copy all the cel values and styles
        self.__copy_sheet_attributes(self.__active_sheet, target_file.active_sheet)

    def __copy_sheet_attributes(self, source_sheet, target_sheet):
        from copy import copy

        target_sheet.sheet_format = copy(source_sheet.sheet_format)
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
        target_sheet.print_options = copy(source_sheet.print_options)
        target_sheet.protection = copy(source_sheet.protection)
        target_sheet.sheet_state = copy(source_sheet.sheet_state)
        target_sheet.views = copy(source_sheet.views)
        target_sheet.data_validations = copy(source_sheet.data_validations)

        # ширина по умолчанию
        if source_sheet.sheet_format.defaultColWidth is not None:
            target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

        # копируем размеры строк
        for row_idx, row_dim in source_sheet.row_dimensions.items():
            target_sheet.row_dimensions[row_idx].height = row_dim.height
            target_sheet.row_dimensions[row_idx].hidden = row_dim.hidden

        # копируем размеры колонок
        for col, col_dim in source_sheet.column_dimensions.items():
            target_dim = target_sheet.column_dimensions[col]
            target_dim.min = col_dim.min
            target_dim.max = col_dim.max
            target_dim.width = col_dim.width
            target_dim.hidden = col_dim.hidden

    def __copy_cells(self, source_sheet, target_sheet):

        for (row, col), source_cell in source_sheet._cells.items():
            if isinstance(source_cell, MergedCell):
                continue

            target_cell = target_sheet.cell(row=row, column=col)
            target_cell.value = source_cell.value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

        # объединённые ячейки копируем ТОЛЬКО здесь
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))


